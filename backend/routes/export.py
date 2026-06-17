"""
Export and notification routes.

Endpoints:
- GET /api/export/budget/{municipality_id}/{month}/csv - Export to CSV
- GET /api/export/budget/{municipality_id}/history/csv - Export history to CSV
- GET /api/export/runs/csv - Export all runs to CSV
- POST /api/notifications/subscribe - Subscribe to email notifications
- POST /api/notifications/unsubscribe - Unsubscribe from notifications
"""

from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from io import StringIO, BytesIO
import csv
from typing import Optional
from datetime import datetime
from collections import defaultdict
import re

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from backend.database import get_db
from backend.models import Municipality, MonthlyRun, BudgetLine, User
from backend.utils.auth_guards import require_login, require_admin, require_municipality_access

router = APIRouter(
    prefix="/api",
    tags=["export-notifications"],
)

_MONTH_PATTERN = re.compile(r"^\d{4}-\d{2}$")
_HEADER_BLUE = '1E3A5F'
_HEADER_FILL = PatternFill(fill_type='solid', fgColor=_HEADER_BLUE)
_HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF')
_BODY_FONT = Font(name='Arial')
_BOLD_FONT = Font(name='Arial', bold=True)
_RTL_ALIGNMENT = Alignment(horizontal='right', vertical='center')


def _validate_month_format(month: str) -> None:
    if not _MONTH_PATTERN.match(month):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail='month must be in YYYY-MM format',
        )


def _normalize_topic_code(code: str) -> str:
    """Normalize topic code so 03 and 3 are treated the same."""
    s = (code or '').strip()
    if not s:
        return ''
    return s.lstrip('0') or '0'


def _style_sheet_headers(ws, header_row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _RTL_ALIGNMENT


def _apply_body_style(ws, start_row: int, end_row: int, col_count: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = _BODY_FONT
            cell.alignment = _RTL_ALIGNMENT


def _autosize_columns(ws, max_col: int) -> None:
    for col in range(1, max_col + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            text_len = len(str(val)) if val is not None else 0
            max_len = max(max_len, text_len)
        ws.column_dimensions[chr(64 + col)].width = min(max(max_len + 2, 12), 40)


@router.get('/export/excel/{month}')
def export_month_summary_to_excel(
    month: str,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Export all municipality runs for a month into a styled Hebrew Excel workbook."""
    _validate_month_format(month)

    runs = db.query(MonthlyRun).filter(MonthlyRun.month == month).all()
    if not runs:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail='No runs found for this month',
        )

    run_ids = [r.id for r in runs]
    municipality_ids = sorted({r.municipality_id for r in runs})

    municipalities = db.query(Municipality).filter(Municipality.id.in_(municipality_ids)).all()
    municipality_by_id = {m.id: m for m in municipalities}

    lines = db.query(BudgetLine).filter(BudgetLine.run_id.in_(run_ids)).all()
    lines_by_run = defaultdict(list)
    for line in lines:
        lines_by_run[line.run_id].append(line)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = 'סיכום'
    ws_summary.sheet_view.rightToLeft = True

    summary_headers = [
        'שם רשות',
        'קוד רשות',
        'קוד 3 (גני ילדים)',
        'קוד 19 (עוזרות)',
        'קוד 33 (גננות מדינה)',
        'קוד 5/45 (קבס)',
        'סך רטרו',
        'סך הכל',
        'סטטוס',
    ]
    ws_summary.append(summary_headers)
    _style_sheet_headers(ws_summary, header_row=1, col_count=len(summary_headers))

    grand_c3 = 0.0
    grand_c19 = 0.0
    grand_c33 = 0.0
    grand_c5_45 = 0.0
    grand_retro = 0.0
    grand_total = 0.0

    for run in sorted(runs, key=lambda r: municipality_by_id.get(r.municipality_id).name if municipality_by_id.get(r.municipality_id) else ''):
        municipality = municipality_by_id.get(run.municipality_id)
        if not municipality:
            continue

        run_lines = lines_by_run.get(run.id, [])
        code_sums = defaultdict(float)
        retro_total = 0.0
        total = 0.0

        for line in run_lines:
            amount = float(line.amount or 0)
            normalized_code = _normalize_topic_code(line.topic_code)
            code_sums[normalized_code] += amount
            total += amount
            if line.is_retro:
                retro_total += amount

        c3 = code_sums.get('3', 0.0)
        c19 = code_sums.get('19', 0.0)
        c33 = code_sums.get('33', 0.0)
        c5_45 = code_sums.get('5', 0.0) + code_sums.get('45', 0.0)
        status_text = 'מאוזן' if run.is_balanced else 'חריגה'

        ws_summary.append([
            municipality.name,
            municipality.code,
            c3,
            c19,
            c33,
            c5_45,
            retro_total,
            total,
            status_text,
        ])

        grand_c3 += c3
        grand_c19 += c19
        grand_c33 += c33
        grand_c5_45 += c5_45
        grand_retro += retro_total
        grand_total += total

    total_row_idx = ws_summary.max_row + 1
    ws_summary.append([
        'סה"כ',
        '',
        grand_c3,
        grand_c19,
        grand_c33,
        grand_c5_45,
        grand_retro,
        grand_total,
        '',
    ])

    _apply_body_style(ws_summary, start_row=2, end_row=ws_summary.max_row, col_count=len(summary_headers))
    for col in range(1, len(summary_headers) + 1):
        ws_summary.cell(row=total_row_idx, column=col).font = _BOLD_FONT

    for row in range(2, ws_summary.max_row + 1):
        for col in [3, 4, 5, 6, 7, 8]:
            ws_summary.cell(row=row, column=col).number_format = '#,##0.00'

    _autosize_columns(ws_summary, len(summary_headers))

    ws_detail = wb.create_sheet(title='פירוט')
    ws_detail.sheet_view.rightToLeft = True
    detail_headers = ['שם רשות', 'קוד נושא', 'תאור נושא', 'סכום', 'חודש תחולה', 'האם רטרו']
    ws_detail.append(detail_headers)
    _style_sheet_headers(ws_detail, header_row=1, col_count=len(detail_headers))

    for run in runs:
        municipality = municipality_by_id.get(run.municipality_id)
        if not municipality:
            continue
        for line in lines_by_run.get(run.id, []):
            ws_detail.append([
                municipality.name,
                line.topic_code,
                line.budget_topic,
                float(line.amount or 0),
                line.period_month,
                'כן' if line.is_retro else 'לא',
            ])

    if ws_detail.max_row > 1:
        _apply_body_style(ws_detail, start_row=2, end_row=ws_detail.max_row, col_count=len(detail_headers))
        for row in range(2, ws_detail.max_row + 1):
            ws_detail.cell(row=row, column=4).number_format = '#,##0.00'

    _autosize_columns(ws_detail, len(detail_headers))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'summary_{month}.xlsx'
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )


# ========== CSV EXPORT ENDPOINTS ==========

@router.get("/export/budget/{municipality_id}/{month}/csv")
def export_budget_to_csv(
    municipality_id: int,
    month: str,
    current_user: User = Depends(require_login),
    db: Session = Depends(get_db)
):
    """
    Export budget data for a specific month to CSV format.
    
    **Requires authentication.** Municipality users can only export their own data.
    
    Args:
        municipality_id: ID of municipality
        month: Month in YYYY-MM format
        db: Database session
        
    Returns:
        CSV file as attachment
    """
    # Check municipality access
    require_municipality_access(municipality_id, current_user)
    
    # Get municipality
    municipality = db.query(Municipality).filter(
        Municipality.id == municipality_id
    ).first()
    
    if not municipality:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Municipality not found"
        )
    
    # Get monthly run
    run = db.query(MonthlyRun).filter(
        MonthlyRun.municipality_id == municipality_id,
        MonthlyRun.month == month,
    ).first()
    
    if not run:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No budget data for this month"
        )
    
    # Get budget lines
    budget_lines = db.query(BudgetLine).filter(
        BudgetLine.run_id == run.id
    ).all()
    
    # Create CSV in memory
    output = StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow([
        "Municipality",
        "Municipality Code",
        "Month",
        "Budget Topic",
        "Topic Code",
        "Amount",
        "Period Month",
        "Type",
        "Is Retro",
        "Notes",
    ])
    
    # Data rows
    for line in budget_lines:
        writer.writerow([
            municipality.name,
            municipality.code,
            month,
            line.budget_topic,
            line.topic_code,
            line.amount,
            line.period_month,
            line.line_type,
            "Yes" if line.is_retro else "No",
            line.notes or "",
        ])
    
    # Add summary
    writer.writerow([])
    writer.writerow(["SUMMARY"])
    writer.writerow(["Invoice Total", run.invoice_total])
    writer.writerow(["Breakdown Total", run.breakdown_total])
    writer.writerow(["Balanced", "Yes" if run.is_balanced else "No"])
    writer.writerow(["Difference", run.difference])
    
    csv_content = output.getvalue()
    
    filename = f"budget_{municipality.code}_{month}.csv"
    
    return StreamingResponse(
        iter([csv_content.encode('utf-8-sig')]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/budget/{municipality_id}/history/csv")
def export_budget_history_to_csv(
    municipality_id: int,
    months: int = 3,
    current_user: User = Depends(require_login),
    db: Session = Depends(get_db)
):
    """
    Export budget history (last N months) to CSV format.
    
    **Requires authentication.** Municipality users can only export their own data.
    
    Args:
        municipality_id: ID of municipality
        months: Number of months to export (default 3, max 12)
        db: Database session
        
    Returns:
        CSV file as attachment
    """
    # Check municipality access
    require_municipality_access(municipality_id, current_user)
    
    # Get municipality
    municipality = db.query(Municipality).filter(
        Municipality.id == municipality_id
    ).first()
    
    if not municipality:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Municipality not found"
        )
    
    # Get monthly runs
    runs = db.query(MonthlyRun).filter(
        MonthlyRun.municipality_id == municipality_id
    ).order_by(MonthlyRun.month.desc()).limit(min(months, 12)).all()
    
    # Create CSV in memory
    output = StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow([
        "Municipality",
        "Month",
        "Budget Topic",
        "Amount",
        "Type",
        "Is Retro",
    ])
    
    # Data rows
    for run in runs:
        budget_lines = db.query(BudgetLine).filter(
            BudgetLine.run_id == run.id
        ).all()
        
        for line in budget_lines:
            writer.writerow([
                municipality.name,
                run.month,
                line.budget_topic,
                line.amount,
                line.line_type,
                "Yes" if line.is_retro else "No",
            ])
    
    csv_content = output.getvalue()
    
    filename = f"budget_history_{municipality.code}.csv"
    
    return StreamingResponse(
        iter([csv_content.encode('utf-8-sig')]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/runs/csv")
def export_runs_to_csv(
    municipality_id: Optional[int] = None,
    current_user: User = Depends(require_login),
    db: Session = Depends(get_db)
):
    """
    Export monthly runs (upload history) to CSV.
    
    **Requires authentication.** Admins can export all runs; municipality users only their own.
    
    Args:
        municipality_id: Optional filter by municipality
        db: Database session
        
    Returns:
        CSV file as attachment
    """
    # Get runs
    query = db.query(MonthlyRun)
    
    # If municipality user without explicit municipality_id, filter by their municipality
    if current_user.role == "municipality":
        if municipality_id and municipality_id != current_user.municipality_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You do not have access to other municipalities' runs"
            )
        if not municipality_id:
            municipality_id = current_user.municipality_id
    
    if municipality_id:
        query = query.filter(MonthlyRun.municipality_id == municipality_id)
    
    runs = query.order_by(MonthlyRun.uploaded_at.desc()).all()
    
    # Create CSV in memory
    output = StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow([
        "Month",
        "Municipality ID",
        "Invoice Total",
        "Breakdown Total",
        "Balanced",
        "Difference",
        "Status",
        "Uploaded At",
        "File Name",
    ])
    
    # Data rows
    for run in runs:
        writer.writerow([
            run.month,
            run.municipality_id,
            run.invoice_total,
            run.breakdown_total,
            "Yes" if run.is_balanced else "No",
            run.difference,
            run.status,
            run.uploaded_at.isoformat(),
            run.file_name,
        ])
    
    csv_content = output.getvalue()
    
    filename = f"monthly_runs_{datetime.now().strftime('%Y%m%d')}.csv"
    
    return StreamingResponse(
        iter([csv_content.encode('utf-8-sig')]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ========== EMAIL NOTIFICATION ENDPOINTS ==========

from pydantic import BaseModel, EmailStr, Field


class EmailNotificationRequest(BaseModel):
    """Request to subscribe/unsubscribe from notifications."""
    email: EmailStr
    municipality_id: Optional[int] = None
    notify_on_upload: bool = True
    notify_on_anomaly: bool = True


@router.post("/notifications/subscribe")
def subscribe_to_notifications(
    subscription: EmailNotificationRequest,
    db: Session = Depends(get_db)
):
    """
    Subscribe to email notifications.
    
    After implementation, this would:
    - Store email subscription preferences
    - Start sending notifications on relevant events
    
    For now, returns a mock success response.
    
    Args:
        subscription: Email and notification preferences
        db: Database session
        
    Returns:
        Success message
    """
    return {
        "status": "success",
        "message": "Subscription created successfully",
        "email": subscription.email,
        "notifications": {
            "on_upload": subscription.notify_on_upload,
            "on_anomaly": subscription.notify_on_anomaly,
        }
    }


@router.post("/notifications/unsubscribe")
def unsubscribe_from_notifications(
    email: EmailStr,
    db: Session = Depends(get_db)
):
    """
    Unsubscribe from email notifications.
    
    Args:
        email: Email to unsubscribe
        db: Database session
        
    Returns:
        Success message
    """
    return {
        "status": "success",
        "message": f"Unsubscribed {email} from notifications"
    }


@router.get("/notifications/preferences/{email}")
def get_notification_preferences(
    email: EmailStr,
    db: Session = Depends(get_db)
):
    """
    Get notification preferences for an email.
    
    Args:
        email: Email address
        db: Database session
        
    Returns:
        Current notification preferences
    """
    return {
        "email": email,
        "subscribed": True,
        "notifications": {
            "on_upload": True,
            "on_anomaly": True,
            "summary_weekly": True,
        }
    }
